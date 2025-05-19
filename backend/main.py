# Імпортування необхідних бібліотек
import io
import json
import os
from datetime import datetime, timedelta
from typing import List, Optional

import numpy as np
import pandas as pd
from bson import ObjectId
from dotenv import load_dotenv
from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.security import OAuth2PasswordBearer
from jose import JWTError, jwt
from motor.motor_asyncio import AsyncIOMotorClient
from passlib.context import CryptContext
from pydantic import BaseModel, validator

from models.turnover_model import TurnoverPredictor

# Завантаження змінних середовища з .env файлу
load_dotenv()

# Ініціалізація FastAPI застосунку
app = FastAPI(title="Employee Turnover Predictor API")

# Визначення OAuth2 схеми для авторизації
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="api/auth/login")

# Завантаження моделі прогнозування плинності персоналу
predictor = TurnoverPredictor()
model = predictor.model
scaler = predictor.scaler
department_encoding = predictor.department_encoding

# Додавання middleware для підтримки CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:3001"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Підключення до MongoDB
MONGODB_URL = os.getenv("MONGODB_URL", "mongodb://admin:password123@localhost:27017")
client = AsyncIOMotorClient(MONGODB_URL)
db = client.turnover_predictor

# Контекст для хешування паролів
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Конфігурація JWT
SECRET_KEY = os.getenv("SECRET_KEY", "your-super-secret-key-here")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 1440

# Модель користувача для автентифікації
class User(BaseModel):
    username: str
    password: str

# Модель для токену автентифікації
class Token(BaseModel):
    access_token: str
    token_type: str

# Дані токену (наприклад, username)
class TokenData(BaseModel):
    username: Optional[str] = None

# Модель працівника з валідацією дат
class Employee(BaseModel):
    name: str
    department: str
    position: str
    risk_level: str
    turnover_probability: float
    email: str
    hire_date: datetime
    salary: float
    performance_score: float
    last_evaluation_date: datetime
    projects: list[str]
    skills: list[str]
    age: Optional[int] = None
    years_of_experience: Optional[int] = None
    work_hours: Optional[int] = None
    projects_completed: Optional[int] = None
    training_hours: Optional[int] = None

    # Обробка строкових форматів дат
    @validator('hire_date', 'last_evaluation_date', pre=True)
    def parse_dates(cls, value):
        if isinstance(value, str):
            try:
                return datetime.fromisoformat(value)
            except ValueError:
                try:
                    return datetime.fromisoformat(value.split('T')[0])
                except ValueError:
                    raise ValueError(f"Invalid date format: {value}")
        return value

# Перевірка відповідності паролю
def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

# Хешування паролю
def get_password_hash(password):
    return pwd_context.hash(password)

# Створення JWT токену
def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

# Отримання поточного користувача за токеном
async def get_current_user(token: str = Depends(oauth2_scheme)):
    print(f"Attempting to authenticate user with token: {token[:10]}...")
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        print(f"Decoded token payload, username: {username}")
        if username is None:
            print("Username not found in token payload")
            raise credentials_exception
        token_data = TokenData(username=username)
    except JWTError as e:
        print(f"JWT decode error: {str(e)}")
        raise credentials_exception
    user = await db.users.find_one({"username": token_data.username})
    print(f"Found user in database: {user is not None}")
    if user is None:
        print(f"User {token_data.username} not found in database")
        raise credentials_exception
    return user

# Реєстрація нового користувача
@app.post("/api/auth/register", response_model=Token)
async def register(user: User):
    if await db.users.find_one({"username": user.username}):
        raise HTTPException(
            status_code=400,
            detail="Username already registered"
        )
    hashed_password = get_password_hash(user.password)
    user_dict = user.dict()
    user_dict["password"] = hashed_password
    await db.users.insert_one(user_dict)

    access_token = create_access_token(
        data={"sub": user.username}
    )
    return {"access_token": access_token, "token_type": "bearer"}

# Вхід користувача
@app.post("/api/auth/login", response_model=Token)
async def login(user: User):
    print(f"Attempting login for user: {user.username}")
    db_user = await db.users.find_one({"username": user.username})
    print(f"Found user in DB: {db_user is not None}")

    if not db_user:
        print("User not found in database")
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )

    if not verify_password(user.password, db_user["password"]):
        print("Password verification failed")
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )

    print("Login successful")
    access_token = create_access_token(
        data={"sub": user.username}
    )
    return {"access_token": access_token, "token_type": "bearer"}

# Клас для серіалізації об'єктів, які не підтримуються стандартним JSON (наприклад, ObjectId, datetime)
class JSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, ObjectId):
            return str(obj)
        if isinstance(obj, datetime):
            return obj.isoformat()
        return super().default(obj)

# Отримання статистики по працівниках
@app.get("/api/employees/stats", response_model=dict)
async def get_employee_stats(current_user: dict = Depends(get_current_user)):
    try:
        # Отримання всіх працівників з бази даних
        employees = await db.employees.find().to_list(length=None)

        # Ініціалізація підрахунків ризику
        risk_counts = {"High Risk": 0, "Medium Risk": 0, "Low Risk": 0}
        total_risk = 0

        # Розрахунок ризиків для кожного працівника
        for employee in employees:
            dept_encoding = department_encoding.get(employee["department"], 1)
            features = np.array([[
                employee.get("age", 30),
                employee.get("years_of_experience", 5),
                employee["salary"],
                dept_encoding,
                employee["performance_score"],
                employee.get("work_hours", 40),
                employee.get("projects_completed", len(employee.get("projects", []))),
                employee.get("training_hours", 20),
                employee.get("time_since_promotion", 12),
                employee.get("team_size", 8)
            ]])
            features_scaled = scaler.transform(features)
            probability = float(model.predict_proba(features_scaled)[0][1])
            total_risk += probability

            # Класифікація ризику
            if probability < 0.3:
                risk_counts["Low Risk"] += 1
            elif probability < 0.7:
                risk_counts["Medium Risk"] += 1
            else:
                risk_counts["High Risk"] += 1

        # Підрахунок статистики по департаментам
        departments = {}
        for employee in employees:
            dept = employee["department"]
            if dept not in departments:
                departments[dept] = {"count": 0, "total_risk": 0}
            departments[dept]["count"] += 1
            dept_encoding = department_encoding.get(dept, 1)
            features = np.array([[
                employee.get("age", 30),
                employee.get("years_of_experience", 5),
                employee["salary"],
                dept_encoding,
                employee["performance_score"],
                employee.get("work_hours", 40),
                employee.get("projects_completed", len(employee.get("projects", []))),
                employee.get("training_hours", 20),
                employee.get("time_since_promotion", 12),
                employee.get("team_size", 8)
            ]])
            features_scaled = scaler.transform(features)
            probability = float(model.predict_proba(features_scaled)[0][1])
            departments[dept]["total_risk"] += probability

        # Формування розподілу по департаментам
        department_distribution = [
            {
                "department": dept,
                "count": data["count"],
                "average_risk": data["total_risk"] / data["count"]
            }
            for dept, data in departments.items()
        ]

        # Обчислення загальної статистики
        total_employees = len(employees)
        average_risk = total_risk / total_employees if total_employees > 0 else 0

        return {
            "total_employees": total_employees,
            "high_risk_count": risk_counts["High Risk"],
            "medium_risk_count": risk_counts["Medium Risk"],
            "low_risk_count": risk_counts["Low Risk"],
            "average_risk": average_risk,
            "department_count": len(departments),
            "department_distribution": department_distribution,
        }
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to fetch employee stats: {str(e)}"
        )



@app.get("/api/employees/top-risk")
async def get_top_risk_employees(
        current_user: dict = Depends(get_current_user),
        limit: int = 5
):
    try:

        employees = await db.employees.find().to_list(length=None)


        for employee in employees:

            dept_encoding = department_encoding.get(employee["department"], 1)


            features = np.array([[
                employee.get("age", 30),
                employee.get("years_of_experience", 5),
                employee["salary"],
                dept_encoding,
                employee["performance_score"],
                employee.get("work_hours", 40),
                employee.get("projects_completed", len(employee.get("projects", []))),
                employee.get("training_hours", 20),
                employee.get("time_since_promotion", 12),
                employee.get("team_size", 8)
            ]])


            features_scaled = scaler.transform(features)


            probability = float(model.predict_proba(features_scaled)[0][1])


            employee["_id"] = str(employee["_id"])
            employee["turnover_probability"] = probability


            if probability < 0.3:
                employee["risk_level"] = "Low Risk"
            elif probability < 0.7:
                employee["risk_level"] = "Medium Risk"
            else:
                employee["risk_level"] = "High Risk"


            if "hire_date" in employee:
                employee["hire_date"] = employee["hire_date"].isoformat()
            if "last_evaluation_date" in employee:
                employee["last_evaluation_date"] = employee["last_evaluation_date"].isoformat()


        employees.sort(key=lambda x: x["turnover_probability"], reverse=True)
        top_risk_employees = employees[:limit]

        return json.loads(json.dumps(top_risk_employees, cls=JSONEncoder))
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to fetch top risk employees: {str(e)}"
        )


@app.get("/api/employees/department-risks")
async def get_department_risk_distribution(
        current_user: dict = Depends(get_current_user)
):
    try:

        employees = await db.employees.find().to_list(length=None)


        departments = {}


        for employee in employees:
            dept = employee["department"]
            if dept not in departments:
                departments[dept] = {
                    "_id": dept,
                    "risk_distribution": [
                        {"risk_level": "High Risk", "count": 0},
                        {"risk_level": "Medium Risk", "count": 0},
                        {"risk_level": "Low Risk", "count": 0}
                    ]
                }


            dept_encoding = department_encoding.get(dept, 1)
            features = np.array([[
                employee.get("age", 30),
                employee.get("years_of_experience", 5),
                employee["salary"],
                dept_encoding,
                employee["performance_score"],
                employee.get("work_hours", 40),
                employee.get("projects_completed", len(employee.get("projects", []))),
                employee.get("training_hours", 20),
                employee.get("time_since_promotion", 12),
                employee.get("team_size", 8)
            ]])

            features_scaled = scaler.transform(features)
            probability = float(model.predict_proba(features_scaled)[0][1])


            risk_index = 2
            if probability >= 0.7:
                risk_index = 0
            elif probability >= 0.3:
                risk_index = 1

            departments[dept]["risk_distribution"][risk_index]["count"] += 1

        return list(departments.values())
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to get department risk distribution: {str(e)}"
        )


@app.get("/api/employees")
async def get_employees(
        current_user: dict = Depends(get_current_user),
        search: Optional[str] = None,
        risk_level: Optional[str] = None,
        sort_by: Optional[str] = "name",
        sort_order: Optional[str] = "asc"
):
    try:
        query = {}
        if search:
            query["$or"] = [
                {"name": {"$regex": search, "$options": "i"}},
                {"first_name": {"$regex": search, "$options": "i"}},
                {"last_name": {"$regex": search, "$options": "i"}},
                {"department": {"$regex": search, "$options": "i"}},
                {"position": {"$regex": search, "$options": "i"}},
                {"email": {"$regex": search, "$options": "i"}}
            ]
        if risk_level and risk_level != "all":

            risk_map = {
                "low": "Low Risk",
                "medium": "Medium Risk",
                "high": "High Risk"
            }
            query["risk_level"] = risk_map.get(risk_level.lower(), risk_level)


        sort_direction = 1 if sort_order == "asc" else -1
        sort_field = sort_by if sort_by in ["name", "department", "position", "risk_level", "salary"] else "name"

        employees = await db.employees.find(query).sort(sort_field, sort_direction).to_list(length=100)


        for employee in employees:

            dept_encoding = department_encoding.get(employee["department"], 1)


            features = np.array([[
                employee.get("age", 30),
                employee.get("years_of_experience", 5),
                employee["salary"],
                dept_encoding,
                employee["performance_score"],
                employee.get("work_hours", 40),
                employee.get("projects_completed", len(employee.get("projects", []))),
                employee.get("training_hours", 20),
                employee.get("time_since_promotion", 12),
                employee.get("team_size", 8)
            ]])


            features_scaled = scaler.transform(features)


            probability = model.predict_proba(features_scaled)[0][1]


            employee["_id"] = str(employee["_id"])
            employee["turnover_probability"] = float(probability)


            if probability < 0.3:
                employee["risk_level"] = "Low Risk"
            elif probability < 0.7:
                employee["risk_level"] = "Medium Risk"
            else:
                employee["risk_level"] = "High Risk"


            if "hire_date" in employee:
                employee["hire_date"] = employee["hire_date"].isoformat()
            if "last_evaluation_date" in employee:
                employee["last_evaluation_date"] = employee["last_evaluation_date"].isoformat()

        return employees
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to fetch employees: {str(e)}"
        )


@app.get("/api/employees/{employee_id}")
async def get_employee(
        employee_id: str,
        current_user: dict = Depends(get_current_user)
):
    try:

        employee = await db.employees.find_one({"_id": ObjectId(employee_id)})
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")

        employee["_id"] = str(employee["_id"])
        return employee
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to fetch employee: {str(e)}"
        )


@app.post("/api/employees")
async def create_employee(
        employee: Employee,
        current_user: dict = Depends(get_current_user)
):
    try:

        existing_employee = await db.employees.find_one({"email": employee.email})
        if existing_employee:
            raise HTTPException(
                status_code=400,
                detail="Employee with this email already exists"
            )

        employee_dict = employee.dict()
        result = await db.employees.insert_one(employee_dict)
        return {"id": str(result.inserted_id), "message": "Employee created successfully"}
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to create employee: {str(e)}"
        )


@app.put("/api/employees/{employee_id}")
async def update_employee(
        employee_id: str,
        employee: Employee,
        current_user: dict = Depends(get_current_user)
):
    try:

        existing_employee = await db.employees.find_one({"_id": ObjectId(employee_id)})
        if not existing_employee:
            raise HTTPException(
                status_code=404,
                detail="Employee not found"
            )


        if employee.email != existing_employee["email"]:
            email_exists = await db.employees.find_one({"email": employee.email})
            if email_exists:
                raise HTTPException(
                    status_code=400,
                    detail="Email already in use by another employee"
                )


        employee_dict = employee.dict()
        result = await db.employees.update_one(
            {"_id": ObjectId(employee_id)},
            {"$set": employee_dict}
        )

        if result.modified_count == 0:
            raise HTTPException(
                status_code=400,
                detail="No changes were made"
            )

        return {"message": "Employee updated successfully"}
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to update employee: {str(e)}"
        )


@app.delete("/api/employees/{employee_id}")
async def delete_employee(
        employee_id: str,
        current_user: dict = Depends(get_current_user)
):
    try:
        result = await db.employees.delete_one({"_id": ObjectId(employee_id)})
        if result.deleted_count == 0:
            raise HTTPException(
                status_code=404,
                detail="Employee not found"
            )
        return {"message": "Employee deleted successfully"}
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to delete employee: {str(e)}"
        )


@app.post("/predict")
async def predict_turnover(employee_data: Employee):
    try:

        dept_encoding = department_encoding.get(employee_data.department, 1)


        features = np.array([[
            employee_data.age or 30,
            employee_data.years_of_experience or 5,
            employee_data.salary,
            dept_encoding,
            employee_data.performance_score,
            employee_data.work_hours or 40,
            employee_data.projects_completed or len(employee_data.projects),
            employee_data.training_hours or 20,
            12,
            8
        ]])


        features_scaled = scaler.transform(features)


        probability = model.predict_proba(features_scaled)[0][1]


        if probability < 0.3:
            risk_level = "Low Risk"
        elif probability < 0.7:
            risk_level = "Medium Risk"
        else:
            risk_level = "High Risk"

        return {
            "turnover_probability": float(probability),
            "risk_level": risk_level
        }
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Prediction failed: {str(e)}"
        )


@app.post("/predict-batch")
async def predict_batch(employees: List[Employee]):
    try:
        predictions = []
        for employee in employees:
            dept_encoding = department_encoding.get(employee.department, 1)
            features = np.array([[
                employee.age or 30,
                employee.years_of_experience or 5,
                employee.salary,
                dept_encoding,
                employee.performance_score,
                employee.work_hours or 40,
                employee.projects_completed or len(employee.projects),
                employee.training_hours or 20,
                12,
                8
            ]])
            features_scaled = scaler.transform(features)
            prediction = model.predict_proba(features_scaled)[0][1]
            risk_factors = []
            if prediction < 0.3:
                risk_level = "Low Risk"
            elif prediction < 0.7:
                risk_level = "Medium Risk"
            else:
                risk_level = "High Risk"
            predictions.append({
                "turnover_probability": float(prediction),
                "risk_level": risk_level
            })
        return predictions
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/export")
async def export_data(
        format: str = Query(..., enum=['pdf', 'excel']),
        current_user: dict = Depends(get_current_user)
):
    try:

        cursor = db.employees.find({})
        employees = []
        async for doc in cursor:
            doc["_id"] = str(doc["_id"])
            employees.append(doc)

        if not employees:
            raise HTTPException(
                status_code=404,
                detail="No employees found to export"
            )

        df = pd.DataFrame(employees)


        columns_to_export = {
            "name": "Name",
            "department": "Department",
            "position": "Position",
            "email": "Email",
            "salary": "Salary",
            "performance_score": "Performance Score",
            "risk_level": "Risk Level",
            "turnover_probability": "Turnover Probability",
            "hire_date": "Hire Date",
        }


        export_df = df[list(columns_to_export.keys())].copy()

        export_df.rename(columns=columns_to_export, inplace=True)


        export_df["Turnover Probability"] = export_df["Turnover Probability"].apply(
            lambda x: f"{x * 100:.1f}%" if pd.notnull(x) else "N/A"
        )
        export_df["Salary"] = export_df["Salary"].apply(
            lambda x: f"${int(x):,}" if pd.notnull(x) else "N/A"
        )
        export_df["Hire Date"] = pd.to_datetime(export_df["Hire Date"]).dt.strftime('%Y-%m-%d')

        if format == 'pdf':
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
            from reportlab.lib.styles import getSampleStyleSheet


            output = io.BytesIO()


            doc = SimpleDocTemplate(output, pagesize=letter)
            elements = []


            styles = getSampleStyleSheet()
            title = Paragraph("Employee Turnover Risk Report", styles['Heading1'])
            elements.append(title)


            data = [list(export_df.columns)]
            data.extend(export_df.values.tolist())


            table = Table(data)


            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ])


            for i, row in enumerate(data[1:], 1):
                risk_index = list(columns_to_export.keys()).index("risk_level")
                risk_level = row[risk_index].lower() if row[risk_index] else ""

                if "high" in risk_level:
                    style.add('BACKGROUND', (risk_index, i), (risk_index, i), colors.lightcoral)
                elif "medium" in risk_level:
                    style.add('BACKGROUND', (risk_index, i), (risk_index, i), colors.lightyellow)
                elif "low" in risk_level:
                    style.add('BACKGROUND', (risk_index, i), (risk_index, i), colors.lightgreen)

            table.setStyle(style)
            elements.append(table)


            doc.build(elements)
            output.seek(0)


            today = datetime.now().strftime('%Y-%m-%d')
            filename = f"employee_turnover_report_{today}.pdf"

            return StreamingResponse(
                output,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        else:

            output = io.BytesIO()


            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                export_df.to_excel(writer, sheet_name="Employee Data", index=False)


                workbook = writer.book
                worksheet = writer.sheets["Employee Data"]


                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

                for col_num, column in enumerate(export_df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')


                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[column_letter].width = adjusted_width


                from openpyxl.formatting.rule import CellIsRule


                risk_col = None
                for i, col in enumerate(export_df.columns):
                    if col == "Risk Level":
                        risk_col = i + 1

                if risk_col:

                    worksheet.conditional_formatting.add(
                        f"{chr(64 + risk_col)}2:{chr(64 + risk_col)}{len(export_df) + 1}",
                        CellIsRule(
                            operator="containsText",
                            formula=['"High"'],
                            stopIfTrue=True,
                            fill=PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        )
                    )
                    worksheet.conditional_formatting.add(
                        f"{chr(64 + risk_col)}2:{chr(64 + risk_col)}{len(export_df) + 1}",
                        CellIsRule(
                            operator="containsText",
                            formula=['"Medium"'],
                            stopIfTrue=True,
                            fill=PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                        )
                    )
                    worksheet.conditional_formatting.add(
                        f"{chr(64 + risk_col)}2:{chr(64 + risk_col)}{len(export_df) + 1}",
                        CellIsRule(
                            operator="containsText",
                            formula=['"Low"'],
                            stopIfTrue=True,
                            fill=PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                        )
                    )

            output.seek(0)


            today = datetime.now().strftime('%Y-%m-%d')
            filename = f"employee_turnover_report_{today}.xlsx"

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class EmployeeData:
    """Helper class to make employee data compatible with the prediction model"""

    def __init__(self, data):
        self.age = data.get("age", 30)
        self.years_of_experience = data.get("years_of_experience", 5)
        self.salary = data.get("salary", 25000)
        self.department = data.get("department", "Engineering")
        self.performance_rating = data.get("performance_score", 3.0)
        self.work_hours = data.get("work_hours", 40)
        self.projects_completed = data.get("projects_completed", 3)
        self.training_hours = data.get("training_hours", 10)
        self.time_since_promotion = data.get("time_since_promotion", 12)
        self.team_size = data.get("team_size", 8)


@app.post("/api/retention-recommendations")
async def get_retention_recommendations(
        request_data: dict,
        current_user: dict = Depends(get_current_user)
):
    try:
        employee_id = request_data.get("employee_id")
        if not employee_id:
            raise HTTPException(
                status_code=400,
                detail="Employee ID is required"
            )


        employee = await db.employees.find_one({"_id": ObjectId(employee_id)})
        if not employee:
            raise HTTPException(
                status_code=404,
                detail="Employee not found"
            )


        baseline_employee = employee.copy()


        original_values = request_data.get("original_values")
        if original_values:
            print(f"Using original values as baseline: {original_values}")

            for key, value in original_values.items():
                if key in baseline_employee and value is not None:

                    if key in ["salary", "performance_score", "training_hours"]:
                        try:
                            baseline_employee[key] = float(value)
                        except (ValueError, TypeError):
                            print(f"Failed to convert {key} value '{value}' to float")
                    elif key in ["work_hours", "years_of_experience"]:
                        try:
                            baseline_employee[key] = int(value)
                        except (ValueError, TypeError):
                            print(f"Failed to convert {key} value '{value}' to int")
                    else:
                        baseline_employee[key] = value


        current_values = request_data.get("current_values")
        if current_values:
            print(f"Received current values: {current_values}")

            for key, value in current_values.items():
                if key in employee:

                    if key in ["salary", "performance_score", "training_hours"] and value is not None:
                        try:
                            employee[key] = float(value)
                        except (ValueError, TypeError):
                            print(f"Failed to convert {key} value '{value}' to float")
                    elif key in ["work_hours", "years_of_experience"] and value is not None:
                        try:
                            employee[key] = int(value)
                        except (ValueError, TypeError):
                            print(f"Failed to convert {key} value '{value}' to int")
                    else:
                        employee[key] = value

            print(
                f"Employee data after updates: {employee['salary']}, {employee['performance_score']}, {employee['work_hours']}, {employee['training_hours']}")


        baseline_employee_data = EmployeeData(baseline_employee)


        baseline_prediction = predictor.predict(baseline_employee_data)
        baseline_probability = baseline_prediction['turnover_probability']
        baseline_risk_level = baseline_prediction['risk_level'].title() + " Risk"


        employee_data = EmployeeData(employee)


        result = predictor.predict_with_feature_impact(employee_data)


        feature_impacts = result.get('feature_impacts', [])
        print(f"Generated {len(feature_impacts)} feature impacts")


        recommendations = []
        for impact in feature_impacts:

            if impact['impact'] > 0:
                continue

            recommendations.append({
                "factor": impact['feature'].replace('_', ' ').title(),
                "current_value": impact['current_value'],
                "recommended_value": impact['recommended_value'],
                "estimated_impact": impact['impact'],
                "action": impact['action'],
                "explanation": get_explanation_for_feature(impact['feature'])
            })


        recommendations.sort(key=lambda x: x['estimated_impact'])
        print(f"Returning {len(recommendations)} recommendations")


        current_prediction = result['prediction']
        current_probability = current_prediction['turnover_probability']
        current_risk_level = current_prediction['risk_level'].title() + " Risk"


        total_impact = sum(rec["estimated_impact"] for rec in recommendations)
        new_probability = max(0.05, current_probability + total_impact)


        if new_probability < 0.3:
            new_risk_level = "Low Risk"
        elif new_probability < 0.6:
            new_risk_level = "Medium Risk"
        else:
            new_risk_level = "High Risk"

        response_data = {
            "employee_id": employee_id,
            "current_probability": baseline_probability if original_values else current_probability,
            "current_risk_level": baseline_risk_level if original_values else current_risk_level,
            "estimated_new_probability": new_probability,
            "estimated_new_risk_level": new_risk_level,
            "recommendations": recommendations,
            "message": "ML-based recommendations to reduce turnover risk."
        }

        return response_data
    except Exception as e:
        print(f"Error in retention recommendations: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Failed to generate recommendations: {str(e)}"
        )


def get_explanation_for_feature(feature):
    """Return detailed explanation for each feature"""
    explanations = {
        "age": "Age is a demographic factor that cannot be changed but may influence retention strategies.",
        "salary": "Compensation is consistently one of the top factors in employee retention decisions.",
        "performance_rating": "Performance scores often correlate with engagement and job satisfaction.",
        "work_hours": "Work-life balance is crucial for preventing burnout and improving retention.",
        "training_hours": "Professional development opportunities improve skills and increase employee loyalty."
    }
    return explanations.get(feature, "This factor has been identified by our model as impacting retention.")


@app.get("/health")
async def health_check():
    return {"status": "healthy"}


@app.get("/export-with-recommendations")
async def export_data_with_recommendations(
        format: str = Query(..., enum=['pdf', 'excel']),
        current_user: dict = Depends(get_current_user)
):
    try:

        cursor = db.employees.find({})
        employees = []
        async for doc in cursor:
            doc["_id"] = str(doc["_id"])
            employees.append(doc)

        if not employees:
            raise HTTPException(
                status_code=404,
                detail="No employees found to export"
            )


        enhanced_employees = []

        for employee in employees:

            employee_data = EmployeeData(employee)
            result = predictor.predict_with_feature_impact(employee_data)


            base_prediction = result['prediction']
            recommendations = []


            for impact in result.get('feature_impacts', []):

                if impact['impact'] > 0:
                    continue

                recommendations.append({
                    "factor": impact['feature'].replace('_', ' ').title(),
                    "current_value": impact['current_value'],
                    "recommended_value": impact['recommended_value'],
                    "estimated_impact": impact['impact'],
                    "action": impact['action'],
                    "explanation": get_explanation_for_feature(impact['feature'])
                })


            recommendations.sort(key=lambda x: x['estimated_impact'])


            recommendation_text = ""
            for i, rec in enumerate(recommendations[:3], 1):
                impact_percent = abs(rec['estimated_impact'] * 100)
                recommendation_text += f"{i}. {rec['action']} (Impact: {impact_percent:.1f}%)\n"


            enhanced_employee = {
                "Name": employee.get("name", ""),
                "Department": employee.get("department", ""),
                "Position": employee.get("position", ""),
                "Email": employee.get("email", ""),
                "Salary": employee.get("salary", 0),
                "Performance Score": employee.get("performance_score", 0),
                "Turnover Probability": base_prediction['turnover_probability'],
                "Risk Level": base_prediction['risk_level'].title() + " Risk",
                "Top Recommendations": recommendation_text.strip(),
                "Hire Date": employee.get("hire_date", "")
            }

            enhanced_employees.append(enhanced_employee)


        df = pd.DataFrame(enhanced_employees)


        df["Turnover Probability"] = df["Turnover Probability"].apply(
            lambda x: f"{x * 100:.1f}%" if pd.notnull(x) else "N/A"
        )
        df["Salary"] = df["Salary"].apply(
            lambda x: f"${int(x):,}" if pd.notnull(x) else "N/A"
        )
        if "Hire Date" in df.columns:
            df["Hire Date"] = pd.to_datetime(df["Hire Date"]).dt.strftime('%Y-%m-%d')

        if format == 'pdf':
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import landscape, letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch


            output = io.BytesIO()


            doc = SimpleDocTemplate(output, pagesize=landscape(letter))
            elements = []


            styles = getSampleStyleSheet()
            title_style = styles['Heading1']
            subtitle_style = styles['Heading2']
            normal_style = styles['Normal']


            recommendation_style = ParagraphStyle(
                'RecommendationStyle',
                parent=styles['BodyText'],
                fontName='Helvetica',
                fontSize=8,
                leading=10,
                spaceAfter=6
            )


            title = Paragraph("Employee Turnover Risk Assessment with Recommendations", title_style)
            elements.append(title)

            today = datetime.now().strftime('%Y-%m-%d')
            date_text = Paragraph(f"Generated on: {today}", normal_style)
            elements.append(date_text)
            elements.append(Spacer(1, 0.25 * inch))



            table_cols = ["Name", "Department", "Position", "Salary", "Performance Score",
                          "Turnover Probability", "Risk Level"]


            header_row = [Paragraph(col, styles["Heading2"]) for col in table_cols]
            header_row.append(Paragraph("Top Recommendations", styles["Heading2"]))


            data = [header_row]

            for _, row in df.iterrows():
                table_row = [Paragraph(str(row[col]), normal_style) for col in table_cols]


                recommendation_text = row["Top Recommendations"]
                recommendation_text = recommendation_text.replace("\n", "<br/>")
                recommendation_para = Paragraph(recommendation_text, recommendation_style)

                table_row.append(recommendation_para)
                data.append(table_row)


            table = Table(data, repeatRows=1)


            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'TOP')
            ])


            risk_index = table_cols.index("Risk Level")
            for i, row in enumerate(data[1:], 1):
                risk_text = str(row[risk_index].text).lower()

                if "high" in risk_text:
                    table_style.add('BACKGROUND', (risk_index, i), (risk_index, i), colors.lightcoral)
                elif "medium" in risk_text:
                    table_style.add('BACKGROUND', (risk_index, i), (risk_index, i), colors.lightyellow)
                elif "low" in risk_text:
                    table_style.add('BACKGROUND', (risk_index, i), (risk_index, i), colors.lightgreen)

            table.setStyle(table_style)
            elements.append(table)


            doc.build(elements)
            output.seek(0)


            today = datetime.now().strftime('%Y-%m-%d')
            filename = f"employee_retention_recommendations_{today}.pdf"

            return StreamingResponse(
                output,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        else:

            output = io.BytesIO()


            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name="Employee Recommendations", index=False)


                workbook = writer.book
                worksheet = writer.sheets["Employee Recommendations"]


                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

                for col_num, column in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')


                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[column_letter].width = adjusted_width


                from openpyxl.formatting.rule import CellIsRule


                risk_col = None
                for i, col in enumerate(df.columns):
                    if col == "Risk Level":
                        risk_col = i + 1

                if risk_col:

                    worksheet.conditional_formatting.add(
                        f"{chr(64 + risk_col)}2:{chr(64 + risk_col)}{len(df) + 1}",
                        CellIsRule(
                            operator="containsText",
                            formula=['"High"'],
                            stopIfTrue=True,
                            fill=PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        )
                    )
                    worksheet.conditional_formatting.add(
                        f"{chr(64 + risk_col)}2:{chr(64 + risk_col)}{len(df) + 1}",
                        CellIsRule(
                            operator="containsText",
                            formula=['"Medium"'],
                            stopIfTrue=True,
                            fill=PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                        )
                    )
                    worksheet.conditional_formatting.add(
                        f"{chr(64 + risk_col)}2:{chr(64 + risk_col)}{len(df) + 1}",
                        CellIsRule(
                            operator="containsText",
                            formula=['"Low"'],
                            stopIfTrue=True,
                            fill=PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                        )
                    )


                detailed_recommendations = []

                for employee in employees:
                    employee_data = EmployeeData(employee)
                    result = predictor.predict_with_feature_impact(employee_data)

                    for impact in result.get('feature_impacts', []):
                        if impact['impact'] > 0:
                            continue

                        detailed_recommendations.append({
                            "Employee Name": employee.get("name", ""),
                            "Department": employee.get("department", ""),
                            "Risk Level": result['prediction']['risk_level'].title() + " Risk",
                            "Factor": impact['feature'].replace('_', ' ').title(),
                            "Current Value": impact['current_display'],
                            "Recommended Value": impact['improved_display'],
                            "Impact %": f"{abs(impact['impact'] * 100):.1f}%",
                            "Action": impact['action'],
                            "Explanation": impact['explanation']
                        })

                if detailed_recommendations:
                    detailed_df = pd.DataFrame(detailed_recommendations)
                    detailed_df.to_excel(writer, sheet_name="Detailed Recommendations", index=False)


                    detailed_ws = writer.sheets["Detailed Recommendations"]


                    for col_num, column in enumerate(detailed_df.columns, 1):
                        cell = detailed_ws.cell(row=1, column=col_num)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center')


                    for column in detailed_ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2) * 1.2
                        detailed_ws.column_dimensions[column_letter].width = adjusted_width

            output.seek(0)


            today = datetime.now().strftime('%Y-%m-%d')
            filename = f"employee_retention_recommendations_{today}.xlsx"

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
    except Exception as e:
        print(f"Export error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn


    print("\n=== Registered Routes ===")
    for route in app.routes:
        print(f"{route.methods} {route.path}")
    print("=======================\n")

    uvicorn.run(app, host="0.0.0.0", port=5000)
